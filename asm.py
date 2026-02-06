#!/usr/bin/env python3
"""
ASM — External Attack Surface Management automation script.

Orchestrates subfinder, amass, nmap, httpx, nuclei, and cloud asset discovery
to perform recurring recon. Preserves full scan history, diffs between runs,
and exports XLSX reports.

Usage:
    ./asm.py scan   -t TARGET [-c CONFIG] [-d DATA_DIR] [--skip MODULES] ...
    ./asm.py diff    -t TARGET --scan-a TS --scan-b TS
    ./asm.py report  -t TARGET --scan TS [--format xlsx|csv]
    ./asm.py list-scans -t TARGET
    ./asm.py --self-test
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import os
import platform
import re
import shutil
import socket
import subprocess
import sys
import tempfile
import textwrap
import time
import urllib.error
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Optional

import yaml

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import boto3

    HAS_BOTO3 = True
except ImportError:
    HAS_BOTO3 = False

_PROJECT_DIR = Path(__file__).resolve().parent
_LOCAL_BIN = _PROJECT_DIR / "bin"
if _LOCAL_BIN.is_dir():
    os.environ["PATH"] = str(_LOCAL_BIN) + os.pathsep + os.environ.get("PATH", "")

# ─── Constants ───────────────────────────────────────────────────────────────

VERSION = "1.0.0"
DEFAULT_CONFIG_NAME = "asm_config.yaml"
SCAN_MODULES = ["subdomains", "ports", "http", "vulns", "cloud"]
TIMESTAMP_FMT = "%Y%m%d_%H%M%S"

EXIT_SUCCESS = 0
EXIT_FATAL = 1
EXIT_PARTIAL = 2
EXIT_NO_TOOLS = 3

LOG_FMT = "%(asctime)s [%(levelname)-7s] %(message)s"
LOG_DATE_FMT = "%Y-%m-%d %H:%M:%S"

# ─── Logging ─────────────────────────────────────────────────────────────────

logger = logging.getLogger("asm")


def setup_logging(verbose: bool = False, log_file: Optional[Path] = None) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logger.setLevel(level)
    logger.handlers.clear()

    console = logging.StreamHandler(sys.stderr)
    console.setLevel(level)
    console.setFormatter(logging.Formatter(LOG_FMT, LOG_DATE_FMT))
    logger.addHandler(console)

    if log_file:
        fh = logging.FileHandler(str(log_file))
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(logging.Formatter(LOG_FMT, LOG_DATE_FMT))
        logger.addHandler(fh)


# ─── Config ──────────────────────────────────────────────────────────────────

DEFAULTS: dict[str, Any] = {
    "tools": {
        "subfinder": "subfinder",
        "amass": "amass",
        "nmap": "nmap",
        "httpx": "httpx",
        "nuclei": "nuclei",
        "cloud_enum": "cloud_enum",
    },
    "scan": {
        "ports": "21-23,25,53,80,110,111,135,139,143,443,445,993,995,"
        "1723,3306,3389,5432,5900,8080,8443,8888,9090",
        "severity": "low,medium,high,critical",
        "timeout": 600,
        "httpx_threads": 50,
        "nuclei_rate_limit": 150,
    },
    "data_dir": "data",
    "cloud": {
        "http_probe": True,
        "permutations": [
            "", "-dev", "-staging", "-prod", "-backup",
            "-assets", "-static", "-media", "-logs", "-data",
        ],
        "aws_region": "us-east-1",
    },
    "report": {
        "format": "xlsx",
        "tls_expiry_warn_days": 30,
        "risky_ports": [21, 23, 3389, 5900, 445, 135, 139],
    },
}


class Config:
    def __init__(self, path: Optional[str] = None):
        self._data = _deep_copy(DEFAULTS)
        if path:
            self._load(path)
        else:
            default = Path(__file__).resolve().parent / DEFAULT_CONFIG_NAME
            if default.is_file():
                self._load(str(default))

    def _load(self, path: str) -> None:
        try:
            with open(path, "r") as f:
                user = yaml.safe_load(f) or {}
            _deep_merge(self._data, user)
            logger.debug("Loaded config from %s", path)
        except Exception as exc:
            logger.warning("Failed to load config %s: %s", path, exc)

    def get(self, *keys: str, default: Any = None) -> Any:
        node = self._data
        for k in keys:
            if isinstance(node, dict) and k in node:
                node = node[k]
            else:
                return default
        return node

    def hash(self) -> str:
        return hashlib.sha256(
            json.dumps(self._data, sort_keys=True).encode()
        ).hexdigest()[:12]


def _deep_copy(d: dict) -> dict:
    return json.loads(json.dumps(d))


def _deep_merge(base: dict, override: dict) -> None:
    for k, v in override.items():
        if k in base and isinstance(base[k], dict) and isinstance(v, dict):
            _deep_merge(base[k], v)
        else:
            base[k] = v


# ─── Tool Runner ─────────────────────────────────────────────────────────────


@dataclass
class ToolResult:
    tool: str
    command: list[str]
    returncode: int
    stdout: str
    stderr: str
    timed_out: bool
    duration: float
    output_file: Optional[str] = None


def check_tool_available(name: str) -> bool:
    return shutil.which(name) is not None


def get_tool_version(name: str) -> str:
    if not check_tool_available(name):
        return "not found"
    for flag in ["-version", "--version", "-V"]:
        try:
            r = subprocess.run(
                [name, flag],
                capture_output=True, text=True, timeout=10,
            )
            out = (r.stdout + r.stderr).strip()
            if out:
                return out.split("\n")[0][:100]
        except Exception:
            continue
    return "unknown"


def run_tool(
    cmd: list[str],
    tool_name: str = "",
    timeout: int = 600,
    cwd: Optional[str] = None,
    output_file: Optional[str] = None,
) -> ToolResult:
    tool_name = tool_name or cmd[0]
    logger.info("Running: %s", " ".join(cmd))
    start = time.time()
    timed_out = False
    try:
        proc = subprocess.run(
            cmd,
            capture_output=True, text=True,
            timeout=timeout, cwd=cwd,
        )
        rc = proc.returncode
        stdout, stderr = proc.stdout, proc.stderr
    except subprocess.TimeoutExpired as exc:
        timed_out = True
        rc = -1
        stdout = exc.stdout.decode("utf-8", errors="replace") if exc.stdout else ""
        stderr = exc.stderr.decode("utf-8", errors="replace") if exc.stderr else ""
        logger.warning("Tool %s timed out after %ds", tool_name, timeout)
    except FileNotFoundError:
        logger.warning("Tool %s not found", tool_name)
        return ToolResult(
            tool=tool_name, command=cmd, returncode=-127,
            stdout="", stderr=f"{tool_name}: command not found",
            timed_out=False, duration=0.0,
        )
    duration = time.time() - start

    if rc != 0 and not timed_out:
        logger.warning("Tool %s exited with code %d", tool_name, rc)
        if stderr:
            logger.debug("stderr: %s", stderr[:500])

    # Read output file if specified and exists
    file_content = None
    if output_file and os.path.isfile(output_file):
        file_content = output_file

    return ToolResult(
        tool=tool_name, command=cmd, returncode=rc,
        stdout=stdout, stderr=stderr,
        timed_out=timed_out, duration=duration,
        output_file=file_content,
    )


def parse_jsonl(text: str) -> list[dict]:
    results = []
    for i, line in enumerate(text.strip().splitlines(), 1):
        line = line.strip()
        if not line:
            continue
        try:
            results.append(json.loads(line))
        except json.JSONDecodeError:
            logger.debug("Skipping malformed JSONL line %d", i)
    return results


# ─── Scan Modules ────────────────────────────────────────────────────────────


def _is_amass_v5(amass_bin: str) -> bool:
    """Detect amass v5+ by checking version output."""
    try:
        r = subprocess.run(
            [amass_bin, "-version"], capture_output=True, text=True, timeout=10,
        )
        version_str = (r.stdout + r.stderr).strip()
        match = re.search(r"v?(\d+)\.", version_str)
        if match:
            return int(match.group(1)) >= 5
    except Exception:
        pass
    return False


def scan_subdomains(
    target: str, cfg: Config, scan_dir: Path, timeout: int,
) -> dict[str, Any]:
    """Subdomain enumeration: subfinder + amass + crt.sh, merged."""
    all_subs: dict[str, set[str]] = {}  # hostname -> set of sources
    cert_entries: list[dict] = []

    # --- subfinder ---
    subfinder_bin = cfg.get("tools", "subfinder", default="subfinder")
    if check_tool_available(subfinder_bin):
        out_file = str(scan_dir / "raw_subfinder.json")
        result = run_tool(
            [subfinder_bin, "-d", target, "-all", "-recursive", "-oJ", "-o", out_file],
            tool_name="subfinder", timeout=timeout,
        )
        raw = ""
        if os.path.isfile(out_file):
            with open(out_file) as f:
                raw = f.read()
        if not raw:
            raw = result.stdout
        for rec in parse_jsonl(raw):
            host = rec.get("host", "").lower().strip()
            if host:
                all_subs.setdefault(host, set()).add("subfinder")
    else:
        logger.warning("subfinder not found, skipping")

    # --- amass ---
    amass_bin = cfg.get("tools", "amass", default="amass")
    if check_tool_available(amass_bin):
        amass_v5 = _is_amass_v5(amass_bin)
        if amass_v5:
            # amass v5: enum -d TARGET writes plain hostnames to stdout
            out_file = str(scan_dir / "raw_amass.txt")
            result = run_tool(
                [amass_bin, "enum", "-d", target, "-o", out_file],
                tool_name="amass", timeout=timeout,
            )
            raw = ""
            if os.path.isfile(out_file):
                with open(out_file) as f:
                    raw = f.read()
            if not raw:
                raw = result.stdout
            for line in raw.strip().splitlines():
                name = line.strip().lower()
                if name and not name.startswith("#"):
                    all_subs.setdefault(name, set()).add("amass")
        else:
            # amass v3/v4: enum -passive -d TARGET -json FILE
            out_file = str(scan_dir / "raw_amass.json")
            result = run_tool(
                [amass_bin, "enum", "-passive", "-d", target, "-json", out_file],
                tool_name="amass", timeout=timeout,
            )
            raw = ""
            if os.path.isfile(out_file):
                with open(out_file) as f:
                    raw = f.read()
            if not raw:
                raw = result.stdout
            for rec in parse_jsonl(raw):
                name = rec.get("name", "").lower().strip()
                if name:
                    all_subs.setdefault(name, set()).add("amass")
    else:
        logger.warning("amass not found, skipping")

    # --- crt.sh ---
    try:
        url = f"https://crt.sh/?q=%25.{target}&output=json"
        req = urllib.request.Request(url, headers={"User-Agent": "ASM-Scanner/1.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8", errors="replace"))
        seen_ids = set()
        for entry in data:
            names_raw = entry.get("name_value", "")
            for name in names_raw.split("\n"):
                name = name.strip().lower().lstrip("*.")
                if name and name.endswith(f".{target}") or name == target:
                    all_subs.setdefault(name, set()).add("crt.sh")
            cert_id = entry.get("id")
            if cert_id and cert_id not in seen_ids:
                seen_ids.add(cert_id)
                cert_entries.append({
                    "id": cert_id,
                    "serial_number": entry.get("serial_number", ""),
                    "issuer_name": entry.get("issuer_name", ""),
                    "common_name": entry.get("common_name", ""),
                    "name_value": entry.get("name_value", ""),
                    "not_before": entry.get("not_before", ""),
                    "not_after": entry.get("not_after", ""),
                })
        logger.info("crt.sh returned %d certificate entries", len(cert_entries))
    except Exception as exc:
        logger.warning("crt.sh query failed: %s", exc)

    # Save cert transparency data
    _save_json(scan_dir / "cert_transparency.json", cert_entries)

    # Always include the target domain itself
    all_subs.setdefault(target.lower(), set()).add("target")

    # Build final subdomain list
    subdomains = []
    for hostname in sorted(all_subs):
        subdomains.append({
            "hostname": hostname,
            "sources": sorted(all_subs[hostname]),
            "first_seen": datetime.now(timezone.utc).isoformat(),
        })
    logger.info("Total unique subdomains: %d", len(subdomains))
    return {"subdomains": subdomains, "cert_entries": cert_entries}


def scan_ports(
    target: str, subdomains: list[dict], cfg: Config, scan_dir: Path, timeout: int,
) -> list[dict]:
    """Port scanning: resolve subdomains to IPs, run nmap, parse XML."""
    # Resolve hostnames to unique IPs
    hosts = [s["hostname"] for s in subdomains]
    if not hosts:
        hosts = [target]
    ip_to_hosts: dict[str, list[str]] = {}
    for h in hosts:
        try:
            infos = socket.getaddrinfo(h, None, socket.AF_INET, socket.SOCK_STREAM)
            for info in infos:
                ip = info[4][0]
                ip_to_hosts.setdefault(ip, [])
                if h not in ip_to_hosts[ip]:
                    ip_to_hosts[ip].append(h)
        except socket.gaierror:
            logger.debug("DNS resolution failed for %s", h)
    logger.info("Resolved to %d unique IPs", len(ip_to_hosts))

    if not ip_to_hosts:
        logger.warning("No IPs resolved, skipping port scan")
        return []

    nmap_bin = cfg.get("tools", "nmap", default="nmap")
    if not check_tool_available(nmap_bin):
        logger.warning("nmap not found, skipping port scan")
        return []

    # Write targets file
    targets_file = scan_dir / "nmap_targets.txt"
    targets_file.write_text("\n".join(ip_to_hosts.keys()) + "\n")

    ports = cfg.get("scan", "ports", default=DEFAULTS["scan"]["ports"])
    xml_file = str(scan_dir / "raw_nmap.xml")

    run_tool(
        [nmap_bin, "-sV", "-T4", "--open", "-p", ports,
         "-iL", str(targets_file), "-oX", xml_file],
        tool_name="nmap", timeout=timeout,
    )

    # Parse nmap XML
    port_results = []
    if os.path.isfile(xml_file):
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            for host_el in root.findall(".//host"):
                addr_el = host_el.find("address[@addrtype='ipv4']")
                if addr_el is None:
                    continue
                ip = addr_el.get("addr", "")
                hostnames = ip_to_hosts.get(ip, [])
                for port_el in host_el.findall(".//port"):
                    state_el = port_el.find("state")
                    if state_el is None or state_el.get("state") != "open":
                        continue
                    service_el = port_el.find("service")
                    port_results.append({
                        "ip": ip,
                        "hostnames": hostnames,
                        "port": int(port_el.get("portid", 0)),
                        "protocol": port_el.get("protocol", "tcp"),
                        "service": service_el.get("name", "") if service_el is not None else "",
                        "product": service_el.get("product", "") if service_el is not None else "",
                        "version": service_el.get("version", "") if service_el is not None else "",
                    })
        except ET.ParseError as exc:
            logger.warning("Failed to parse nmap XML: %s", exc)

    logger.info("Found %d open ports", len(port_results))
    return port_results


def scan_http(
    subdomains: list[dict], cfg: Config, scan_dir: Path, timeout: int,
) -> list[dict]:
    """HTTP probing and tech fingerprinting via httpx."""
    httpx_bin = cfg.get("tools", "httpx", default="httpx")
    if not check_tool_available(httpx_bin):
        logger.warning("httpx not found, skipping HTTP probing")
        return []

    hosts = [s["hostname"] for s in subdomains]
    if not hosts:
        logger.warning("No subdomains for HTTP probing")
        return []

    targets_file = scan_dir / "httpx_targets.txt"
    targets_file.write_text("\n".join(hosts) + "\n")

    threads = str(cfg.get("scan", "httpx_threads", default=50))
    out_file = str(scan_dir / "raw_httpx.json")

    run_tool(
        [httpx_bin, "-l", str(targets_file), "-json", "-td", "-sc", "-title",
         "-server", "-cl", "-cdn", "-tls-grab", "-fr", "-threads", threads,
         "-o", out_file],
        tool_name="httpx", timeout=timeout,
    )

    http_results = []
    raw = ""
    if os.path.isfile(out_file):
        with open(out_file) as f:
            raw = f.read()
    for rec in parse_jsonl(raw):
        tls_info = rec.get("tls-grab", rec.get("tls", {})) or {}
        http_results.append({
            "url": rec.get("url", ""),
            "status_code": rec.get("status_code", rec.get("status-code", 0)),
            "title": rec.get("title", ""),
            "web_server": rec.get("webserver", rec.get("server", "")),
            "technologies": rec.get("tech", rec.get("technologies", [])),
            "cdn": rec.get("cdn", False),
            "cdn_name": rec.get("cdn-name", rec.get("cdn_name", "")),
            "content_length": rec.get("content_length", rec.get("content-length", 0)),
            "tls_issuer": tls_info.get("issuer_organization", tls_info.get("issuer", "")),
            "tls_expiry": tls_info.get("not_after", tls_info.get("expiry", "")),
            "tls_subject": tls_info.get("subject_cn", tls_info.get("subject", "")),
        })

    logger.info("Probed %d HTTP services", len(http_results))
    return http_results


def scan_vulns(
    http_services: list[dict], cfg: Config, scan_dir: Path, timeout: int,
) -> list[dict]:
    """Vulnerability scanning via nuclei on live HTTP URLs."""
    nuclei_bin = cfg.get("tools", "nuclei", default="nuclei")
    if not check_tool_available(nuclei_bin):
        logger.warning("nuclei not found, skipping vulnerability scan")
        return []

    urls = [s["url"] for s in http_services if s.get("url")]
    if not urls:
        logger.warning("No live URLs for vulnerability scanning")
        return []

    urls_file = scan_dir / "nuclei_urls.txt"
    urls_file.write_text("\n".join(urls) + "\n")

    severity = cfg.get("scan", "severity", default=DEFAULTS["scan"]["severity"])
    rate_limit = str(cfg.get("scan", "nuclei_rate_limit", default=150))
    out_file = str(scan_dir / "raw_nuclei.json")

    run_tool(
        [nuclei_bin, "-l", str(urls_file), "-jsonl", "-s", severity,
         "-rl", rate_limit, "-as", "-o", out_file],
        tool_name="nuclei", timeout=timeout,
    )

    vuln_results = []
    raw = ""
    if os.path.isfile(out_file):
        with open(out_file) as f:
            raw = f.read()
    for rec in parse_jsonl(raw):
        info = rec.get("info", {})
        vuln_results.append({
            "template_id": rec.get("template-id", rec.get("templateID", "")),
            "name": info.get("name", ""),
            "severity": info.get("severity", "unknown"),
            "host": rec.get("host", ""),
            "matched_at": rec.get("matched-at", rec.get("matched", "")),
            "description": info.get("description", ""),
            "tags": info.get("tags", []),
            "references": info.get("reference", []),
        })

    logger.info("Found %d vulnerabilities", len(vuln_results))
    return vuln_results


def scan_cloud(
    target: str, cfg: Config, scan_dir: Path, timeout: int,
) -> list[dict]:
    """Cloud asset discovery: HTTP probing + optional cloud_enum + boto3."""
    cloud_results: list[dict] = []

    # --- HTTP probing for cloud storage ---
    if cfg.get("cloud", "http_probe", default=True):
        cloud_results.extend(_cloud_http_probe(target, cfg))

    # --- cloud_enum ---
    cloud_enum_bin = cfg.get("tools", "cloud_enum", default="cloud_enum")
    if check_tool_available(cloud_enum_bin):
        out_file = str(scan_dir / "raw_cloud_enum.txt")
        result = run_tool(
            [cloud_enum_bin, "-k", target.split(".")[0], "-l", out_file],
            tool_name="cloud_enum", timeout=timeout,
        )
        if os.path.isfile(out_file):
            with open(out_file) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#"):
                        cloud_results.append({
                            "provider": _guess_provider(line),
                            "type": "storage",
                            "name": line,
                            "url": line,
                            "status": "found",
                            "method": "cloud_enum",
                        })
    else:
        logger.debug("cloud_enum not found, skipping")

    # --- boto3 S3 audit ---
    if HAS_BOTO3:
        cloud_results.extend(_boto3_s3_audit(target, cfg))
    else:
        logger.debug("boto3 not available, skipping S3 ACL audit")

    # Deduplicate by (provider, type, name)
    seen = set()
    unique = []
    for item in cloud_results:
        key = (item.get("provider"), item.get("type"), item.get("name"))
        if key not in seen:
            seen.add(key)
            unique.append(item)
    cloud_results = unique

    logger.info("Found %d cloud assets", len(cloud_results))
    return cloud_results


def _cloud_http_probe(target: str, cfg: Config) -> list[dict]:
    """Probe S3/Azure Blob/GCP Storage endpoints via HTTP."""
    results = []
    base_keywords = _domain_keywords(target)
    permutations = cfg.get("cloud", "permutations", default=[""])

    endpoints: list[tuple[str, str, str]] = []  # (url, provider, bucket_name)
    for kw in base_keywords:
        for perm in permutations:
            name = f"{kw}{perm}"
            endpoints.append(
                (f"https://{name}.s3.amazonaws.com", "aws", name)
            )
            endpoints.append(
                (f"https://{name}.blob.core.windows.net", "azure", name)
            )
            endpoints.append(
                (f"https://storage.googleapis.com/{name}", "gcp", name)
            )

    for url, provider, name in endpoints:
        try:
            req = urllib.request.Request(url, method="HEAD")
            req.add_header("User-Agent", "ASM-Scanner/1.0")
            with urllib.request.urlopen(req, timeout=5) as resp:
                status = resp.getcode()
                if status in (200, 403):
                    label = "public_listable" if status == 200 else "exists_private"
                    results.append({
                        "provider": provider,
                        "type": "storage",
                        "name": name,
                        "url": url,
                        "status": label,
                        "method": "http_probe",
                    })
        except urllib.error.HTTPError as e:
            if e.code == 403:
                results.append({
                    "provider": provider,
                    "type": "storage",
                    "name": name,
                    "url": url,
                    "status": "exists_private",
                    "method": "http_probe",
                })
        except Exception:
            pass
    return results


def _domain_keywords(target: str) -> list[str]:
    """Generate keyword variations from a domain."""
    parts = target.replace(".", "-").split("-")
    keywords = set()
    keywords.add(target.replace(".", "-"))
    keywords.add(target.split(".")[0])
    for p in parts:
        if len(p) > 2:
            keywords.add(p)
    return sorted(keywords)


def _guess_provider(url: str) -> str:
    if "amazonaws" in url or "s3" in url:
        return "aws"
    if "blob.core.windows" in url or "azure" in url:
        return "azure"
    if "storage.googleapis" in url or "gcp" in url:
        return "gcp"
    return "unknown"


def _boto3_s3_audit(target: str, cfg: Config) -> list[dict]:
    """Audit owned S3 buckets for public ACLs/policies using boto3."""
    results = []
    region = cfg.get("cloud", "aws_region", default="us-east-1")
    try:
        s3 = boto3.client("s3", region_name=region)
        response = s3.list_buckets()
        keyword = target.split(".")[0].lower()
        for bucket in response.get("Buckets", []):
            name = bucket["Name"]
            if keyword not in name.lower():
                continue
            status = "private"
            # Check ACL
            try:
                acl = s3.get_bucket_acl(Bucket=name)
                for grant in acl.get("Grants", []):
                    grantee = grant.get("Grantee", {})
                    uri = grantee.get("URI", "")
                    if "AllUsers" in uri or "AuthenticatedUsers" in uri:
                        status = "public_acl"
                        break
            except Exception:
                pass
            # Check bucket policy
            try:
                policy = s3.get_bucket_policy(Bucket=name)
                pol_doc = json.loads(policy["Policy"])
                for stmt in pol_doc.get("Statement", []):
                    principal = stmt.get("Principal", "")
                    if principal == "*" or (
                        isinstance(principal, dict) and principal.get("AWS") == "*"
                    ):
                        status = "public_policy"
                        break
            except s3.exceptions.from_code("NoSuchBucketPolicy"):
                pass
            except Exception:
                pass

            results.append({
                "provider": "aws",
                "type": "s3_bucket",
                "name": name,
                "url": f"https://{name}.s3.amazonaws.com",
                "status": status,
                "method": "boto3_audit",
            })
    except Exception as exc:
        logger.debug("boto3 S3 audit failed: %s", exc)
    return results


# ─── Diff Engine ─────────────────────────────────────────────────────────────


def _diff_by_key(
    current: list[dict],
    previous: list[dict],
    key_func: Callable[[dict], Any],
    compare_fields: Optional[list[str]] = None,
) -> dict[str, list]:
    cur_map = {}
    for item in current:
        k = key_func(item)
        cur_map[k] = item
    prev_map = {}
    for item in previous:
        k = key_func(item)
        prev_map[k] = item

    cur_keys = set(cur_map.keys())
    prev_keys = set(prev_map.keys())

    added = [cur_map[k] for k in sorted(cur_keys - prev_keys, key=str)]
    removed = [prev_map[k] for k in sorted(prev_keys - cur_keys, key=str)]

    changed = []
    if compare_fields:
        for k in sorted(cur_keys & prev_keys, key=str):
            diffs = {}
            for f in compare_fields:
                cv = cur_map[k].get(f)
                pv = prev_map[k].get(f)
                if cv != pv:
                    diffs[f] = {"old": pv, "new": cv}
            if diffs:
                changed.append({"key": str(k), "changes": diffs, "current": cur_map[k]})

    return {"added": added, "removed": removed, "changed": changed}


# Per-module diff definitions
DIFF_DEFS: dict[str, dict[str, Any]] = {
    "subdomains": {
        "key": lambda d: d.get("hostname", ""),
        "compare": None,
    },
    "ports": {
        "key": lambda d: (d.get("ip", ""), d.get("port", 0), d.get("protocol", "")),
        "compare": ["service", "product", "version"],
    },
    "http_services": {
        "key": lambda d: d.get("url", ""),
        "compare": ["status_code", "title", "technologies", "web_server", "tls_expiry"],
    },
    "vulnerabilities": {
        "key": lambda d: (d.get("template_id", ""), d.get("matched_at", "")),
        "compare": None,
    },
    "cloud_assets": {
        "key": lambda d: (d.get("provider", ""), d.get("type", ""), d.get("name", "")),
        "compare": ["status"],
    },
    "cert_transparency": {
        "key": lambda d: d.get("id", ""),
        "compare": None,
    },
}


def compute_diff(current_dir: Path, previous_dir: Path) -> dict[str, Any]:
    """Compute diff between two scan directories."""
    diff_result: dict[str, Any] = {
        "scan_a": previous_dir.name,
        "scan_b": current_dir.name,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "modules": {},
    }

    for module, ddef in DIFF_DEFS.items():
        cur_file = current_dir / f"{module}.json"
        prev_file = previous_dir / f"{module}.json"

        cur_data = _load_json(cur_file) if cur_file.is_file() else []
        prev_data = _load_json(prev_file) if prev_file.is_file() else []

        if not isinstance(cur_data, list):
            cur_data = []
        if not isinstance(prev_data, list):
            prev_data = []

        module_diff = _diff_by_key(cur_data, prev_data, ddef["key"], ddef["compare"])
        module_diff["summary"] = {
            "added": len(module_diff["added"]),
            "removed": len(module_diff["removed"]),
            "changed": len(module_diff["changed"]),
            "current_total": len(cur_data),
            "previous_total": len(prev_data),
        }
        diff_result["modules"][module] = module_diff

    return diff_result


# ─── XLSX Export ─────────────────────────────────────────────────────────────

# Color constants
HDR_FILL = PatternFill("solid", fgColor="2F5496") if HAS_OPENPYXL else None
HDR_FONT = Font(bold=True, color="FFFFFF", size=11) if HAS_OPENPYXL else None
RED_FILL = PatternFill("solid", fgColor="FFCCCC") if HAS_OPENPYXL else None
GREEN_FILL = PatternFill("solid", fgColor="CCFFCC") if HAS_OPENPYXL else None
YELLOW_FILL = PatternFill("solid", fgColor="FFFFCC") if HAS_OPENPYXL else None
ORANGE_FILL = PatternFill("solid", fgColor="FFE0B2") if HAS_OPENPYXL else None
CRIT_FILL = PatternFill("solid", fgColor="FF8A80") if HAS_OPENPYXL else None
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
) if HAS_OPENPYXL else None

SEVERITY_FILLS: dict[str, Any] = {}
if HAS_OPENPYXL:
    SEVERITY_FILLS = {
        "critical": CRIT_FILL,
        "high": RED_FILL,
        "medium": ORANGE_FILL,
        "low": YELLOW_FILL,
        "info": PatternFill("solid", fgColor="E3F2FD"),
    }


def _style_header(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws, col_count: int, max_width: int = 50) -> None:
    for col in range(1, col_count + 1):
        max_len = 0
        letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                try:
                    val = str(cell.value or "")
                    max_len = max(max_len, min(len(val), max_width))
                except Exception:
                    pass
        ws.column_dimensions[letter].width = max(max_len + 2, 10)


def _add_autofilter(ws, col_count: int) -> None:
    ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}{ws.max_row}"


def generate_xlsx(
    scan_dir: Path, report_path: Path, cfg: Config,
    diff_data: Optional[dict] = None,
) -> Path:
    """Generate XLSX report with 7 sheets."""
    if not HAS_OPENPYXL:
        logger.error("openpyxl not installed, cannot generate XLSX report")
        raise RuntimeError("openpyxl is required for XLSX reports")

    wb = Workbook()

    # Load scan data
    meta = _load_json(scan_dir / "scan_meta.json") or {}
    subdomains = _load_json(scan_dir / "subdomains.json") or []
    ports = _load_json(scan_dir / "ports.json") or []
    http_services = _load_json(scan_dir / "http_services.json") or []
    vulns = _load_json(scan_dir / "vulnerabilities.json") or []
    cloud = _load_json(scan_dir / "cloud_assets.json") or []

    risky_ports = set(cfg.get("report", "risky_ports", default=[]))
    tls_warn_days = cfg.get("report", "tls_expiry_warn_days", default=30)

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        ["ASM Scan Report"],
        [],
        ["Target", meta.get("target", "")],
        ["Scan Timestamp", meta.get("start_time", "")],
        ["Duration (s)", meta.get("duration_seconds", "")],
        ["Config Hash", meta.get("config_hash", "")],
        [],
        ["Asset Counts"],
        ["Subdomains", len(subdomains)],
        ["Open Ports", len(ports)],
        ["HTTP Services", len(http_services)],
        ["Vulnerabilities", len(vulns)],
        ["Cloud Assets", len(cloud)],
        [],
        ["Modules Run", ", ".join(meta.get("modules_run", []))],
        ["Modules Skipped", ", ".join(meta.get("modules_skipped", []))],
        ["Modules Failed", ", ".join(meta.get("modules_failed", []))],
    ]
    if diff_data:
        summary_rows.append([])
        summary_rows.append(["Changes Since Previous Scan"])
        for mod, mdata in diff_data.get("modules", {}).items():
            s = mdata.get("summary", {})
            summary_rows.append([
                mod,
                f"+{s.get('added', 0)} added, "
                f"-{s.get('removed', 0)} removed, "
                f"~{s.get('changed', 0)} changed",
            ])

    for row in summary_rows:
        ws.append(row)
    # Style the title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A8"].font = Font(bold=True, size=12)
    _auto_width(ws, 2)

    # --- Sheet 2: Subdomains ---
    ws2 = wb.create_sheet("Subdomains")
    headers = ["Hostname", "Sources", "First Seen"]
    ws2.append(headers)
    _style_header(ws2, len(headers))
    for s in sorted(subdomains, key=lambda x: x.get("hostname", "")):
        sources = s.get("sources", [])
        if isinstance(sources, list):
            sources = ", ".join(sources)
        ws2.append([s.get("hostname", ""), sources, s.get("first_seen", "")])
    _add_autofilter(ws2, len(headers))
    _auto_width(ws2, len(headers))

    # --- Sheet 3: Open Ports ---
    ws3 = wb.create_sheet("Open Ports")
    headers = ["IP", "Hostnames", "Port", "Protocol", "Service", "Product", "Version"]
    ws3.append(headers)
    _style_header(ws3, len(headers))
    for p in ports:
        hostnames = p.get("hostnames", [])
        if isinstance(hostnames, list):
            hostnames = ", ".join(hostnames)
        row_data = [
            p.get("ip", ""), hostnames, p.get("port", ""),
            p.get("protocol", ""), p.get("service", ""),
            p.get("product", ""), p.get("version", ""),
        ]
        ws3.append(row_data)
        if p.get("port") in risky_ports:
            row_num = ws3.max_row
            for col in range(1, len(headers) + 1):
                ws3.cell(row=row_num, column=col).fill = RED_FILL
    _add_autofilter(ws3, len(headers))
    _auto_width(ws3, len(headers))

    # --- Sheet 4: HTTP Services ---
    ws4 = wb.create_sheet("HTTP Services")
    headers = [
        "URL", "Status", "Title", "Server", "Technologies",
        "CDN", "TLS Issuer", "TLS Expiry",
    ]
    ws4.append(headers)
    _style_header(ws4, len(headers))
    for h in http_services:
        tech = h.get("technologies", [])
        if isinstance(tech, list):
            tech = ", ".join(str(t) for t in tech)
        row_data = [
            h.get("url", ""), h.get("status_code", ""),
            h.get("title", ""), h.get("web_server", ""),
            tech, str(h.get("cdn", "")),
            h.get("tls_issuer", ""), h.get("tls_expiry", ""),
        ]
        ws4.append(row_data)
        # Highlight if TLS expiring soon
        tls_exp = h.get("tls_expiry", "")
        if tls_exp and _tls_expiring_soon(tls_exp, tls_warn_days):
            row_num = ws4.max_row
            for col in range(1, len(headers) + 1):
                ws4.cell(row=row_num, column=col).fill = YELLOW_FILL
    _add_autofilter(ws4, len(headers))
    _auto_width(ws4, len(headers))

    # --- Sheet 5: Vulnerabilities ---
    ws5 = wb.create_sheet("Vulnerabilities")
    headers = [
        "Severity", "Template ID", "Name", "Host",
        "Matched At", "Description", "Tags", "References",
    ]
    ws5.append(headers)
    _style_header(ws5, len(headers))
    for v in vulns:
        tags = v.get("tags", [])
        if isinstance(tags, list):
            tags = ", ".join(str(t) for t in tags)
        refs = v.get("references", [])
        if isinstance(refs, list):
            refs = ", ".join(str(r) for r in refs)
        row_data = [
            v.get("severity", ""), v.get("template_id", ""),
            v.get("name", ""), v.get("host", ""),
            v.get("matched_at", ""), v.get("description", ""),
            tags, refs,
        ]
        ws5.append(row_data)
        sev = v.get("severity", "").lower()
        fill = SEVERITY_FILLS.get(sev)
        if fill:
            row_num = ws5.max_row
            for col in range(1, len(headers) + 1):
                ws5.cell(row=row_num, column=col).fill = fill
    _add_autofilter(ws5, len(headers))
    _auto_width(ws5, len(headers))

    # --- Sheet 6: Cloud Assets ---
    ws6 = wb.create_sheet("Cloud Assets")
    headers = ["Provider", "Type", "Name", "URL", "Status", "Method"]
    ws6.append(headers)
    _style_header(ws6, len(headers))
    for c in cloud:
        row_data = [
            c.get("provider", ""), c.get("type", ""),
            c.get("name", ""), c.get("url", ""),
            c.get("status", ""), c.get("method", ""),
        ]
        ws6.append(row_data)
        if c.get("status") == "public_listable":
            row_num = ws6.max_row
            for col in range(1, len(headers) + 1):
                ws6.cell(row=row_num, column=col).fill = RED_FILL
    _add_autofilter(ws6, len(headers))
    _auto_width(ws6, len(headers))

    # --- Sheet 7: Changes ---
    ws7 = wb.create_sheet("Changes")
    headers = ["Module", "Change Type", "Key", "Details"]
    ws7.append(headers)
    _style_header(ws7, len(headers))
    if diff_data:
        for mod, mdata in diff_data.get("modules", {}).items():
            for item in mdata.get("added", []):
                key = _item_key_str(mod, item)
                ws7.append([mod, "added", key, ""])
                ws7.cell(row=ws7.max_row, column=1).fill = GREEN_FILL
                ws7.cell(row=ws7.max_row, column=2).fill = GREEN_FILL
            for item in mdata.get("removed", []):
                key = _item_key_str(mod, item)
                ws7.append([mod, "removed", key, ""])
                ws7.cell(row=ws7.max_row, column=1).fill = RED_FILL
                ws7.cell(row=ws7.max_row, column=2).fill = RED_FILL
            for item in mdata.get("changed", []):
                key = item.get("key", "")
                details = json.dumps(item.get("changes", {}), default=str)
                ws7.append([mod, "changed", key, details])
                ws7.cell(row=ws7.max_row, column=1).fill = YELLOW_FILL
                ws7.cell(row=ws7.max_row, column=2).fill = YELLOW_FILL
    _add_autofilter(ws7, len(headers))
    _auto_width(ws7, len(headers))

    wb.save(str(report_path))
    logger.info("Report saved to %s", report_path)
    return report_path


def generate_csv_report(scan_dir: Path, report_path: Path) -> Path:
    """Fallback CSV export (one file per module)."""
    csv_dir = report_path.parent / report_path.stem
    csv_dir.mkdir(parents=True, exist_ok=True)
    for name in ["subdomains", "ports", "http_services", "vulnerabilities", "cloud_assets"]:
        data = _load_json(scan_dir / f"{name}.json") or []
        if not data:
            continue
        csv_path = csv_dir / f"{name}.csv"
        keys = list(data[0].keys())
        with open(csv_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            for row in data:
                writer.writerow({k: _csv_val(row.get(k, "")) for k in keys})
    logger.info("CSV report saved to %s", csv_dir)
    return csv_dir


def _csv_val(v: Any) -> str:
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    return str(v) if v is not None else ""


def _tls_expiring_soon(expiry_str: str, warn_days: int) -> bool:
    for fmt in ("%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            exp = datetime.strptime(expiry_str, fmt).replace(tzinfo=timezone.utc)
            now = datetime.now(timezone.utc)
            return (exp - now).days < warn_days
        except ValueError:
            continue
    return False


def _item_key_str(module: str, item: dict) -> str:
    ddef = DIFF_DEFS.get(module)
    if ddef:
        try:
            return str(ddef["key"](item))
        except Exception:
            pass
    return str(item.get("hostname", item.get("url", item.get("name", ""))))


# ─── Data Management ────────────────────────────────────────────────────────


def _save_json(path: Path, data: Any) -> None:
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=str)


def _load_json(path: Path) -> Any:
    try:
        with open(path) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return None


def get_target_dir(data_dir: str, target: str) -> Path:
    safe_target = re.sub(r"[^a-zA-Z0-9._-]", "_", target)
    return Path(data_dir) / safe_target


def create_scan_dir(target_dir: Path) -> tuple[Path, str]:
    ts = datetime.now().strftime(TIMESTAMP_FMT)
    scan_dir = target_dir / "scans" / ts
    scan_dir.mkdir(parents=True, exist_ok=True)
    return scan_dir, ts


def update_latest_symlink(target_dir: Path, scan_dir: Path) -> None:
    latest = target_dir / "latest"
    tmp = target_dir / f".latest_tmp_{os.getpid()}"
    try:
        rel_path = os.path.relpath(scan_dir, target_dir)
        os.symlink(rel_path, str(tmp))
        os.rename(str(tmp), str(latest))
    except OSError as exc:
        logger.warning("Failed to update symlink: %s", exc)
        # Cleanup temp if rename failed
        try:
            os.unlink(str(tmp))
        except OSError:
            pass


def find_previous_scan(target_dir: Path, current_ts: str) -> Optional[Path]:
    scans_dir = target_dir / "scans"
    if not scans_dir.is_dir():
        return None
    dirs = sorted(
        [d for d in scans_dir.iterdir() if d.is_dir() and d.name != current_ts],
        key=lambda d: d.name,
        reverse=True,
    )
    return dirs[0] if dirs else None


def list_scans(target_dir: Path) -> list[dict]:
    scans_dir = target_dir / "scans"
    if not scans_dir.is_dir():
        return []
    results = []
    for d in sorted(scans_dir.iterdir(), reverse=True):
        if not d.is_dir():
            continue
        meta = _load_json(d / "scan_meta.json") or {}
        results.append({
            "timestamp": d.name,
            "target": meta.get("target", ""),
            "duration": meta.get("duration_seconds", ""),
            "modules_run": meta.get("modules_run", []),
            "path": str(d),
        })
    return results


# ─── Orchestrator ────────────────────────────────────────────────────────────


def run_full_scan(
    target: str, cfg: Config, data_dir: str,
    skip_modules: list[str], ports_override: Optional[str],
    severity_override: Optional[str], timeout_override: Optional[int],
    generate_report: bool, verbose: bool,
) -> int:
    """Run a full ASM scan: phases 1–7 in sequence."""
    start_time = time.time()
    start_ts = datetime.now(timezone.utc).isoformat()

    target_dir = get_target_dir(data_dir, target)
    scan_dir, ts = create_scan_dir(target_dir)
    reports_dir = target_dir / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)

    # Setup file logging
    setup_logging(verbose, scan_dir / "asm.log")

    logger.info("=" * 60)
    logger.info("ASM Scan — target: %s", target)
    logger.info("Scan directory: %s", scan_dir)
    logger.info("=" * 60)

    # Apply overrides
    if ports_override:
        cfg._data["scan"]["ports"] = ports_override
    if severity_override:
        cfg._data["scan"]["severity"] = severity_override
    timeout = timeout_override or cfg.get("scan", "timeout", default=600)

    modules_run = []
    modules_skipped = []
    modules_failed = []
    errors: list[str] = []
    tool_versions: dict[str, str] = {}

    # Collect tool versions
    for tool_name in ["subfinder", "amass", "nmap", "httpx", "nuclei"]:
        bin_name = cfg.get("tools", tool_name, default=tool_name)
        tool_versions[tool_name] = get_tool_version(bin_name)

    # Phase 1: Subdomain Enumeration
    subdomains = []
    if "subdomains" not in skip_modules:
        logger.info("─── Phase 1: Subdomain Enumeration ───")
        try:
            result = scan_subdomains(target, cfg, scan_dir, timeout)
            subdomains = result["subdomains"]
            _save_json(scan_dir / "subdomains.json", subdomains)
            modules_run.append("subdomains")
        except Exception as exc:
            logger.error("Subdomain enumeration failed: %s", exc)
            modules_failed.append("subdomains")
            errors.append(f"subdomains: {exc}")
    else:
        modules_skipped.append("subdomains")
        # Try to load from previous scan
        prev = find_previous_scan(target_dir, ts)
        if prev:
            prev_subs = _load_json(prev / "subdomains.json")
            if prev_subs:
                subdomains = prev_subs
                logger.info("Loaded %d subdomains from previous scan", len(subdomains))

    # Phase 2: Port Scanning
    port_results = []
    if "ports" not in skip_modules:
        logger.info("─── Phase 2: Port Scanning ───")
        try:
            port_results = scan_ports(target, subdomains, cfg, scan_dir, timeout)
            _save_json(scan_dir / "ports.json", port_results)
            modules_run.append("ports")
        except Exception as exc:
            logger.error("Port scanning failed: %s", exc)
            modules_failed.append("ports")
            errors.append(f"ports: {exc}")
    else:
        modules_skipped.append("ports")

    # Phase 3: HTTP Probing
    http_services = []
    if "http" not in skip_modules:
        logger.info("─── Phase 3: HTTP Probing ───")
        try:
            http_services = scan_http(subdomains, cfg, scan_dir, timeout)
            _save_json(scan_dir / "http_services.json", http_services)
            modules_run.append("http")
        except Exception as exc:
            logger.error("HTTP probing failed: %s", exc)
            modules_failed.append("http")
            errors.append(f"http: {exc}")
    else:
        modules_skipped.append("http")

    # Phase 4: Vulnerability Scanning
    vuln_results = []
    if "vulns" not in skip_modules:
        logger.info("─── Phase 4: Vulnerability Scanning ───")
        try:
            vuln_results = scan_vulns(http_services, cfg, scan_dir, timeout)
            _save_json(scan_dir / "vulnerabilities.json", vuln_results)
            modules_run.append("vulns")
        except Exception as exc:
            logger.error("Vulnerability scanning failed: %s", exc)
            modules_failed.append("vulns")
            errors.append(f"vulns: {exc}")
    else:
        modules_skipped.append("vulns")

    # Phase 5: Cloud Asset Discovery
    cloud_results = []
    if "cloud" not in skip_modules:
        logger.info("─── Phase 5: Cloud Asset Discovery ───")
        try:
            cloud_results = scan_cloud(target, cfg, scan_dir, timeout)
            _save_json(scan_dir / "cloud_assets.json", cloud_results)
            modules_run.append("cloud")
        except Exception as exc:
            logger.error("Cloud discovery failed: %s", exc)
            modules_failed.append("cloud")
            errors.append(f"cloud: {exc}")
    else:
        modules_skipped.append("cloud")

    # Save scan metadata
    duration = time.time() - start_time
    meta = {
        "target": target,
        "start_time": start_ts,
        "end_time": datetime.now(timezone.utc).isoformat(),
        "duration_seconds": round(duration, 2),
        "timestamp": ts,
        "modules_run": modules_run,
        "modules_skipped": modules_skipped,
        "modules_failed": modules_failed,
        "tool_versions": tool_versions,
        "config_hash": cfg.hash(),
        "errors": errors,
        "version": VERSION,
        "counts": {
            "subdomains": len(subdomains),
            "ports": len(port_results),
            "http_services": len(http_services),
            "vulnerabilities": len(vuln_results),
            "cloud_assets": len(cloud_results),
        },
    }
    _save_json(scan_dir / "scan_meta.json", meta)

    # Update symlink
    update_latest_symlink(target_dir, scan_dir)

    # Phase 6: Diff
    diff_data = None
    prev_scan = find_previous_scan(target_dir, ts)
    if prev_scan:
        logger.info("─── Phase 6: Computing Diff ───")
        try:
            diff_data = compute_diff(scan_dir, prev_scan)
            _save_json(reports_dir / f"{ts}_diff.json", diff_data)
            _log_diff_summary(diff_data)
        except Exception as exc:
            logger.error("Diff computation failed: %s", exc)
            errors.append(f"diff: {exc}")
    else:
        logger.info("No previous scan found, skipping diff")

    # Phase 7: Report
    if generate_report:
        logger.info("─── Phase 7: Generating Report ───")
        try:
            report_path = reports_dir / f"{ts}_report.xlsx"
            generate_xlsx(scan_dir, report_path, cfg, diff_data)
        except Exception as exc:
            logger.error("Report generation failed: %s", exc)
            errors.append(f"report: {exc}")

    # Summary
    logger.info("=" * 60)
    logger.info("Scan completed in %.1f seconds", duration)
    logger.info(
        "Results: %d subdomains, %d ports, %d HTTP, %d vulns, %d cloud",
        len(subdomains), len(port_results), len(http_services),
        len(vuln_results), len(cloud_results),
    )
    if modules_failed:
        logger.warning("Failed modules: %s", ", ".join(modules_failed))
    logger.info("=" * 60)

    # Determine exit code
    if modules_failed and not modules_run:
        return EXIT_NO_TOOLS
    if modules_failed:
        return EXIT_PARTIAL
    return EXIT_SUCCESS


def _log_diff_summary(diff_data: dict) -> None:
    for mod, mdata in diff_data.get("modules", {}).items():
        s = mdata.get("summary", {})
        if s.get("added") or s.get("removed") or s.get("changed"):
            logger.info(
                "  %s: +%d added, -%d removed, ~%d changed",
                mod, s["added"], s["removed"], s["changed"],
            )


# ─── Self-test ───────────────────────────────────────────────────────────────


def run_self_test() -> int:
    """Run self-test with inline mock data — no external tools needed."""
    print("ASM Self-Test")
    print("=" * 40)
    errors = []

    # Test 1: Config
    print("[1/7] Config loading...", end=" ")
    try:
        c = Config()
        assert c.get("scan", "timeout") is not None
        assert c.hash()
        print("OK")
    except Exception as exc:
        print(f"FAIL: {exc}")
        errors.append(str(exc))

    # Test 2: JSONL parser
    print("[2/7] JSONL parser...", end=" ")
    try:
        data = parse_jsonl('{"a":1}\n{"b":2}\nbad line\n{"c":3}')
        assert len(data) == 3
        assert data[0]["a"] == 1
        print("OK")
    except Exception as exc:
        print(f"FAIL: {exc}")
        errors.append(str(exc))

    # Test 3: Diff engine
    print("[3/7] Diff engine...", end=" ")
    try:
        prev = [
            {"hostname": "a.example.com"},
            {"hostname": "b.example.com"},
            {"hostname": "c.example.com"},
        ]
        cur = [
            {"hostname": "b.example.com"},
            {"hostname": "c.example.com"},
            {"hostname": "d.example.com"},
        ]
        result = _diff_by_key(cur, prev, lambda d: d["hostname"])
        assert len(result["added"]) == 1
        assert result["added"][0]["hostname"] == "d.example.com"
        assert len(result["removed"]) == 1
        assert result["removed"][0]["hostname"] == "a.example.com"
        assert len(result["changed"]) == 0

        # Test with compare fields
        prev_ports = [{"ip": "1.2.3.4", "port": 80, "protocol": "tcp", "service": "http"}]
        cur_ports = [{"ip": "1.2.3.4", "port": 80, "protocol": "tcp", "service": "https"}]
        result = _diff_by_key(
            cur_ports, prev_ports,
            lambda d: (d["ip"], d["port"], d["protocol"]),
            ["service"],
        )
        assert len(result["changed"]) == 1
        assert result["changed"][0]["changes"]["service"]["old"] == "http"
        assert result["changed"][0]["changes"]["service"]["new"] == "https"
        print("OK")
    except Exception as exc:
        print(f"FAIL: {exc}")
        errors.append(str(exc))

    # Test 4: Data management
    print("[4/7] Data management...", end=" ")
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            td = Path(tmpdir) / "test_target"
            sd, ts = create_scan_dir(td)
            assert sd.is_dir()
            assert ts
            _save_json(sd / "test.json", {"hello": "world"})
            loaded = _load_json(sd / "test.json")
            assert loaded["hello"] == "world"
            update_latest_symlink(td, sd)
            latest = td / "latest"
            assert latest.is_symlink() or latest.exists()
            print("OK")
    except Exception as exc:
        print(f"FAIL: {exc}")
        errors.append(str(exc))

    # Test 5: XLSX generation
    print("[5/7] XLSX generation...", end=" ")
    if HAS_OPENPYXL:
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                sd = Path(tmpdir)
                _save_json(sd / "scan_meta.json", {
                    "target": "test.com", "start_time": "2024-01-01T00:00:00Z",
                    "duration_seconds": 42, "config_hash": "abc123",
                    "modules_run": ["subdomains"], "modules_skipped": [],
                    "modules_failed": [],
                })
                _save_json(sd / "subdomains.json", [
                    {"hostname": "www.test.com", "sources": ["crt.sh"], "first_seen": "2024-01-01"},
                ])
                _save_json(sd / "ports.json", [
                    {"ip": "1.2.3.4", "hostnames": ["www.test.com"],
                     "port": 443, "protocol": "tcp", "service": "https",
                     "product": "nginx", "version": "1.20"},
                ])
                _save_json(sd / "http_services.json", [
                    {"url": "https://www.test.com", "status_code": 200,
                     "title": "Test", "web_server": "nginx",
                     "technologies": ["jQuery"], "cdn": False,
                     "tls_issuer": "Let's Encrypt", "tls_expiry": "2024-06-01"},
                ])
                _save_json(sd / "vulnerabilities.json", [
                    {"template_id": "CVE-2024-1234", "name": "Test Vuln",
                     "severity": "high", "host": "https://www.test.com",
                     "matched_at": "https://www.test.com/path",
                     "description": "Test", "tags": ["cve"], "references": []},
                ])
                _save_json(sd / "cloud_assets.json", [
                    {"provider": "aws", "type": "storage", "name": "test-bucket",
                     "url": "https://test-bucket.s3.amazonaws.com",
                     "status": "exists_private", "method": "http_probe"},
                ])
                report_path = sd / "test_report.xlsx"
                c = Config()
                diff_mock = {
                    "modules": {
                        "subdomains": {
                            "summary": {"added": 1, "removed": 0, "changed": 0},
                            "added": [{"hostname": "new.test.com"}],
                            "removed": [], "changed": [],
                        },
                    },
                }
                generate_xlsx(sd, report_path, c, diff_mock)
                assert report_path.is_file()
                assert report_path.stat().st_size > 0
                print("OK")
        except Exception as exc:
            print(f"FAIL: {exc}")
            errors.append(str(exc))
    else:
        print("SKIP (openpyxl not installed)")

    # Test 6: Domain keywords
    print("[6/7] Domain keywords...", end=" ")
    try:
        kws = _domain_keywords("example.com")
        assert "example" in kws
        assert "example-com" in kws
        print("OK")
    except Exception as exc:
        print(f"FAIL: {exc}")
        errors.append(str(exc))

    # Test 7: Full diff pipeline
    print("[7/7] Full diff pipeline...", end=" ")
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            scan_a = Path(tmpdir) / "scan_a"
            scan_b = Path(tmpdir) / "scan_b"
            scan_a.mkdir()
            scan_b.mkdir()
            _save_json(scan_a / "subdomains.json", [
                {"hostname": "a.test.com"},
                {"hostname": "b.test.com"},
            ])
            _save_json(scan_b / "subdomains.json", [
                {"hostname": "b.test.com"},
                {"hostname": "c.test.com"},
            ])
            _save_json(scan_a / "ports.json", [])
            _save_json(scan_b / "ports.json", [])
            for name in ["http_services", "vulnerabilities", "cloud_assets", "cert_transparency"]:
                _save_json(scan_a / f"{name}.json", [])
                _save_json(scan_b / f"{name}.json", [])

            diff = compute_diff(scan_b, scan_a)
            subs = diff["modules"]["subdomains"]
            assert subs["summary"]["added"] == 1
            assert subs["summary"]["removed"] == 1
            print("OK")
    except Exception as exc:
        print(f"FAIL: {exc}")
        errors.append(str(exc))

    print("=" * 40)
    if errors:
        print(f"FAILED: {len(errors)} test(s)")
        return EXIT_FATAL
    print("All tests passed.")
    return EXIT_SUCCESS


# ─── CLI ─────────────────────────────────────────────────────────────────────


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="asm.py",
        description="ASM — External Attack Surface Management",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""\
            examples:
              ./asm.py scan -t example.com
              ./asm.py scan -t example.com --skip cloud,vulns -v
              ./asm.py diff -t example.com --scan-a 20260101_120000 --scan-b 20260102_120000
              ./asm.py report -t example.com --scan 20260101_120000
              ./asm.py list-scans -t example.com
              ./asm.py --self-test
        """),
    )
    parser.add_argument("--version", action="version", version=f"%(prog)s {VERSION}")
    parser.add_argument("--self-test", action="store_true", help="Run built-in self-test")

    subparsers = parser.add_subparsers(dest="command")

    # scan
    scan_p = subparsers.add_parser("scan", help="Run a full ASM scan")
    scan_p.add_argument("-t", "--target", required=True, help="Target domain")
    scan_p.add_argument("-c", "--config", default=None, help="Config file path")
    scan_p.add_argument("-d", "--data-dir", default=None, help="Data directory")
    scan_p.add_argument(
        "--skip", default="", help="Comma-separated modules to skip: subdomains,ports,http,vulns,cloud"
    )
    scan_p.add_argument("--ports", default=None, help="Override port specification")
    scan_p.add_argument("--severity", default=None, help="Override nuclei severity")
    scan_p.add_argument("--timeout", type=int, default=None, help="Override tool timeout (seconds)")
    scan_p.add_argument("--no-report", action="store_true", help="Skip report generation")
    scan_p.add_argument("-v", "--verbose", action="store_true", help="Verbose output")

    # diff
    diff_p = subparsers.add_parser("diff", help="Compute diff between two scans")
    diff_p.add_argument("-t", "--target", required=True, help="Target domain")
    diff_p.add_argument("--scan-a", required=True, help="Older scan timestamp")
    diff_p.add_argument("--scan-b", required=True, help="Newer scan timestamp")
    diff_p.add_argument("-c", "--config", default=None, help="Config file path")
    diff_p.add_argument("-d", "--data-dir", default=None, help="Data directory")

    # report
    report_p = subparsers.add_parser("report", help="Generate report from existing scan")
    report_p.add_argument("-t", "--target", required=True, help="Target domain")
    report_p.add_argument("--scan", required=True, dest="scan_ts", help="Scan timestamp")
    report_p.add_argument(
        "--format", default="xlsx", choices=["xlsx", "csv"], help="Report format"
    )
    report_p.add_argument("-c", "--config", default=None, help="Config file path")
    report_p.add_argument("-d", "--data-dir", default=None, help="Data directory")

    # list-scans
    list_p = subparsers.add_parser("list-scans", help="List scan history for a target")
    list_p.add_argument("-t", "--target", required=True, help="Target domain")
    list_p.add_argument("-d", "--data-dir", default=None, help="Data directory")
    list_p.add_argument("-c", "--config", default=None, help="Config file path")

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    if args.self_test:
        return run_self_test()

    if not args.command:
        parser.print_help()
        return EXIT_FATAL

    cfg = Config(getattr(args, "config", None))
    data_dir = getattr(args, "data_dir", None) or cfg.get("data_dir", default="data")
    # Resolve relative to script location
    if not os.path.isabs(data_dir):
        data_dir = str(Path(__file__).resolve().parent / data_dir)

    if args.command == "scan":
        setup_logging(args.verbose)
        skip = [s.strip() for s in args.skip.split(",") if s.strip()]
        for s in skip:
            if s not in SCAN_MODULES:
                logger.error("Unknown module to skip: %s (valid: %s)", s, ", ".join(SCAN_MODULES))
                return EXIT_FATAL
        return run_full_scan(
            target=args.target, cfg=cfg, data_dir=data_dir,
            skip_modules=skip, ports_override=args.ports,
            severity_override=args.severity, timeout_override=args.timeout,
            generate_report=not args.no_report, verbose=args.verbose,
        )

    elif args.command == "diff":
        setup_logging(verbose=False)
        target_dir = get_target_dir(data_dir, args.target)
        scan_a_dir = target_dir / "scans" / args.scan_a
        scan_b_dir = target_dir / "scans" / args.scan_b
        if not scan_a_dir.is_dir():
            logger.error("Scan directory not found: %s", scan_a_dir)
            return EXIT_FATAL
        if not scan_b_dir.is_dir():
            logger.error("Scan directory not found: %s", scan_b_dir)
            return EXIT_FATAL
        diff_data = compute_diff(scan_b_dir, scan_a_dir)
        reports_dir = target_dir / "reports"
        reports_dir.mkdir(parents=True, exist_ok=True)
        out_path = reports_dir / f"{args.scan_b}_diff.json"
        _save_json(out_path, diff_data)
        _log_diff_summary(diff_data)
        print(f"Diff saved to {out_path}")
        return EXIT_SUCCESS

    elif args.command == "report":
        setup_logging(verbose=False)
        target_dir = get_target_dir(data_dir, args.target)
        scan_dir = target_dir / "scans" / args.scan_ts
        if not scan_dir.is_dir():
            logger.error("Scan directory not found: %s", scan_dir)
            return EXIT_FATAL
        reports_dir = target_dir / "reports"
        reports_dir.mkdir(parents=True, exist_ok=True)

        # Check for existing diff
        diff_path = reports_dir / f"{args.scan_ts}_diff.json"
        diff_data = _load_json(diff_path) if diff_path.is_file() else None

        if args.format == "xlsx":
            report_path = reports_dir / f"{args.scan_ts}_report.xlsx"
            generate_xlsx(scan_dir, report_path, cfg, diff_data)
        else:
            report_path = reports_dir / f"{args.scan_ts}_report_csv"
            generate_csv_report(scan_dir, report_path)
        print(f"Report saved to {report_path}")
        return EXIT_SUCCESS

    elif args.command == "list-scans":
        target_dir = get_target_dir(data_dir, args.target)
        scans = list_scans(target_dir)
        if not scans:
            print(f"No scans found for {args.target}")
            return EXIT_SUCCESS
        print(f"Scans for {args.target}:")
        print(f"{'Timestamp':<20} {'Duration':>10} {'Modules'}")
        print("-" * 60)
        for s in scans:
            dur = f"{s['duration']}s" if s['duration'] else "?"
            mods = ", ".join(s["modules_run"]) if s["modules_run"] else "?"
            print(f"{s['timestamp']:<20} {dur:>10} {mods}")
        return EXIT_SUCCESS

    parser.print_help()
    return EXIT_FATAL


if __name__ == "__main__":
    sys.exit(main())
